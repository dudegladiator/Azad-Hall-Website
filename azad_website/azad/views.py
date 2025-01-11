from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import *
from .forms import ContactForm, CommentForm, ComplaintForm
from django.core.mail import BadHeaderError, send_mail
from django.http import JsonResponse
from django.core import serializers
from django.contrib.auth.views import LogoutView
from django.contrib.auth import logout
from django.contrib.auth.models import User
from openpyxl import load_workbook
from datetime import datetime, timedelta
from django.contrib import messages
from django.core.paginator import Paginator, Page
from django.views.decorators.csrf import csrf_protect
from datetime import datetime
import requests
import base64
from django.utils.timezone import make_aware

allowedEmailsecratary = True

allowedEmailsNotice = ["arnabdas.9039@gmail.com"]
allowedEmails = [
    "harsh247gupta@gmail.com",
    "harsh90731@gmail.com",
    "rajumeshram767@gmail.com",
    "hariomk628@gmail.com",
    "sg06959.sgsg@gmail.com",
    "pooniakushagra20@gmail.com",
    "somnathmishra1802@gmail.com",
]
allowedEmailsLibrary = [
    "harsh247gupta@gmail.com",
    "pooniakushagra20@gmail.com",
    "harsh90731@gmail.com",
    "rajumeshram767@gmail.com",
    "hariomk628@gmail.com",
    "sg06959.sgsg@gmail.com",
    "pranjalchouhan2014@gmail.com",
    "somnathmishra1802@gmail.com",
]


def importBoardersFromExcel(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb.active
        azad_boarders_to_create = []
        for row in ws.iter_rows(values_only=True):
            azad_boarder = azad_boarders(emails=row[0], books=0)
            azad_boarders_to_create.append(azad_boarder)
        azad_boarders.objects.bulk_create(azad_boarders_to_create)

        return render(request, "index.html")


# For books
def importBooksFromExcel(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb.active
        books_to_create = []

        for row in ws.iter_rows(values_only=True):
            book1 = book(
                title=row[0],
                author=row[1],
                department=row[2],
                shelf=row[3],
                quantity=row[4],
                available=row[4],
            )
            books_to_create.append(book1)
        book.objects.bulk_create(books_to_create)

        return render(request, "index.html")


def alumni(request):
    return render(request, "alumni.html")


def addBooks(request):
    if request.user.is_authenticated:
        email = request.user.email
        if email in allowedEmails:
            return render(request, "addBooks.html")
    messages.info(request, "Please login with valid ID to add books")
    return redirect("/")


def addBoarders(request):
    if request.user.is_authenticated:
        email = request.user.email
        if email in allowedEmails:
            return render(request, "addBoarders.html")
    messages.info(request, "Please login with valid ID to add boarders")
    return redirect("/")


def login(request):
    return render(request, "login.html")


def index(request):
    if request.user.is_authenticated:
        email = request.user.email
        boarders = azad_boarders.objects.all()
        for boarder in boarders:
            if boarder.emails == email:
                name = boarder.name
                params = {"name": name}
                return render(request, "index.html", params)
        logout(request)
        messages.info("Please login with valid EmailID")
        return render(request, "khoj.html")

    return render(request, "index.html")


@csrf_protect
def upload_file_to_imagekit(api_key, file, file_name):
    """
    Upload a file to ImageKit.io V2 API.

    Parameters:
        api_key (str): Your ImageKit private API key.
        file (binary): The file to upload.
        file_name (str): Name of the file to upload.

    Returns:
        dict: API response in JSON format.
    """
    url = "https://upload.imagekit.io/api/v2/files/upload"
    auth_header = base64.b64encode(f"{api_key}:".encode()).decode()

    headers = {
        "Authorization": f"Basic {auth_header}",
    }

    payload = {
        "file": file,
        "fileName": file_name,
    }

    try:
        response = requests.post(url, headers=headers, data=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}


# Function to handle complaints page
def complain(request):
    if request.method == "GET" and request.user.is_authenticated:
        return render(request, "complain.html")
    elif request.method == "POST":
        return HttpResponse("Complaint Submitted Successfully")
    else:
        return HttpResponse(status=405)


@csrf_protect
def profile(request):
    if request.method == "POST" and request.user.is_authenticated:
        boarder = azad_boarders.objects.get(emails=request.user.email)
        if boarder:
            contact = request.POST.get("contact_no")
            name = request.POST.get("name")
            roll_no = request.POST.get("roll_no")
            register = azad_boarders.objects.update(
                roll_no=roll_no,
                name=name,
                contact=contact,
            )
            messages.info(request, "Saved Successfully")
            return redirect("/profile")
        else:
            messages.info(request, "Email not found")

    boarder = azad_boarders.objects.get(emails=request.user.email)
    return render(request, "profile.html", {"user": boarder})


@csrf_protect
def submit_form(request):
    if request.method == "POST" and request.user.is_authenticated:
        category = request.POST.get("category")
        room_no = request.POST.get("room_no")
        complain_text = request.POST.get("complain")
        contact_no = request.POST.get("contact_no")
        image_link = request.POST.get("image")
        upload_image = request.POST.get("file")
        boarder = azad_boarders.objects.get(emails=request.user.email)
        now = datetime.now()
        t_string = now.strftime("%d/%m/%Y %H:%M %p")
        created_at = t_string
        complaints.objects.create(
            name=boarder.name,
            roll_no=boarder.roll_no,
            email=boarder.emails,
            category=category,
            contact_no=contact_no,
            complain=complain_text,
            status="pending",
            room_no=room_no,
            created_at=created_at,
            image_link=image_link,
            review="None",
        )

        messages.info(request, "Complaint submitted successfully")
        return redirect("/")
    else:
        return HttpResponse("Invalid request method")


# Function to display complaints
def showComplaints(request):
    if request.user.is_authenticated:
        email = request.user.email
        if email in allowedEmails:
            pending_complaints = complaints.objects.filter(status="pending")
            pending_paginator = Paginator(pending_complaints, 10)
            current_page_pending = pending_paginator.page(
                request.GET.get("pending_page", 1)
            )
            completed_complaints = complaints.objects.filter(
                status="Completed"
            ).order_by("-id")
            completed_paginator = Paginator(completed_complaints, 10)
            current_page_completed = completed_paginator.page(
                request.GET.get("completed_page", 1)
            )
            ongoing_complaints = complaints.objects.filter(status="Ongoing")
            ongoing_paginator = Paginator(ongoing_complaints, 10)
            current_page_ongoing = ongoing_paginator.page(
                request.GET.get("ongoing_page", 1)
            )
            params = {
                "pending_complaints": Paginator(pending_complaints, 10).get_page(
                    request.GET.get("pending_page", 1)
                ),
                "completed_complaints": Paginator(completed_complaints, 10).get_page(
                    request.GET.get("completed_page", 1)
                ),
                "ongoing_complaints": Paginator(ongoing_complaints, 10).get_page(
                    request.GET.get("ongoing_page", 1)
                ),
            }
            return render(request, "showComplaints.html", params)
        else:
            messages.info(request, "Please login with a valid ID to access complaints")
            return redirect("/")
    return HttpResponse("Unauthorized", status=401)


# Function to display a single complaint
def showFullComplain(request, complain_id):
    if request.user.is_authenticated:
        complaint = complaints.objects.get(id=complain_id)
        return render(request, "fullComplain.html", {"complain": complaint})
    else:
        return HttpResponse("Unauthorized", status=401)


# Function to update complaint status
def updateStatus(request):
    if request.method == "POST":
        id = request.POST.get("id")
        manager_review = request.POST.get("manager_review", "None")
        set_status = request.POST.get("set_status", "Completed")

        complaint = complaints.objects.get(id=id)
        complaint.status = "Ongoing" if set_status == "ongoing" else "Completed"
        complaint.manager_review = manager_review
        complaint.modified_at = datetime.now().strftime("%d/%m/%Y %H:%M %p")
        complaint.save()

        return redirect("/showComplaints")
    return HttpResponse("Invalid request method", status=405)


# Function to display user's complaint status
def complain_status(request):
    if request.user.is_authenticated:
        user_complaints = complaints.objects.filter(email=request.user.email).order_by(
            "-id"
        )
        return render(request, "complain_status.html", {"complains": user_complaints})
    else:
        return HttpResponse("Unauthorized", status=401)


# Function to handle complaint submission with ImageKit integration
def submit_complain(request):
    if request.method == "POST":
        form = ComplaintForm(request.POST, request.FILES)
        if form.is_valid():
            complaint = form.save(commit=False)

            # Handle image upload via ImageKit
            if "upload_image" in request.FILES:
                upload_image = request.FILES["upload_image"]
                api_key = "private_iXnU83xFLdR99tM9YEFMOQLuwus="

                # Read the uploaded file as binary
                response = upload_file_to_imagekit(
                    api_key, upload_image, upload_image.name
                )

                # Check response for success
                if "error" not in response:
                    complaint.image_link = response.get("url")
                    complaint.imagekit_file_id = response.get("fileId")
                else:
                    messages.error(request, f"Image upload failed: {response['error']}")

            # Save complaint
            complaint.save()
            messages.success(request, "Complaint submitted successfully!")
            return redirect("complaint_status")
        else:
            messages.error(request, "Form is invalid. Please correct the errors.")
    else:
        form = ComplaintForm()

    return render(request, "complaint_form.html", {"form": form})


def noticeboard(request):
    if request.user.is_authenticated:
        noticeboard = Notice.objects.all()
        return render(request, "noticeboard.html", {"noticeboard": noticeboard})
    messages.info(request, "Please login with valid ID.")
    return redirect("/")


@csrf_protect
def noticeadd(request):
    if request.method == "POST" and request.user.is_authenticated:
        email = request.user.email
        if email in allowedEmailsNotice:
            title = request.POST.get("title")
            subtitle = request.POST.get("subtitle")
            description = request.POST.get("description")
            image_link = request.POST.get("image")
            event_date_time = request.POST.get("event")
            issue_date_time = datetime.now()
            author = request.POST.get("author")
            Notice.objects.create(
                title=title,
                subtitle=subtitle,
                description=description,
                image=image_link,
                event_date_time=event_date_time,
                issue_date_time=make_aware(issue_date_time),
                author=author,
            )
            users = list(User.objects.all())
            if not users:
                messages.info(request, "Recipients not found")
            emails = [user.email for user in users]
            try:
                send_mail(
                    subject=title,
                    message=description,
                    from_email="arnabdas.9039@gmail.com",
                    recipient_list=emails,
                    fail_silently=False,
                )
                messages.info(request, "Notice posted and emailed successfully")
            except BadHeaderError:
                messages.info(request, "Notice posted successfully")
            return redirect("/noticeboard")
        messages.info(request, "Please login with valid ID.")
        return redirect("/")
    return render(request, "addnotice.html")


def achievements(request):
    achievements = Achievements.objects.all()
    return render(request, "achievements.html", {"achievements": achievements})


def achievement(request, achievementid):
    achievement = Achievements.objects.filter(id=achievementid)
    achievements = Achievements.objects.all()
    return render(
        request,
        "achievement.html",
        {"achievement": achievement[0], "achievements": achievements},
    )


def events(request):
    events = Event.objects.all()
    return render(request, "events.html", {"events": events})


def khoj(request):
    return render(request, "khoj.html")


def library(request, searchedBooks=None, str=None):
    if request.user.is_authenticated:
        form_data = request.session.get("form_data", None)
        if searchedBooks:
            books = searchedBooks
            return render(
                request, "library.html", {"books": books, "searchedString": str}
            )
        else:
            books = book.objects.all().order_by("id")
        books_paginator = Paginator(books, 30)
        current_page_books = books_paginator.page(request.GET.get("books_page", 1))
        context = {
            "books": current_page_books,
            "searchedString": str,
            "form_data": form_data,
        }
        return render(request, "library.html", context)
    messages.info(request, "Please login with valid ID to access library")
    return redirect("/")


def checkout(request):
    if request.method == "POST":
        id = request.POST.get("id")
        Book = book.objects.get(id=id)
        boarder = azad_boarders.objects.get(emails=request.user.email)
        if boarder.books < 2:
            now = datetime.now()
            boarder.books += 1
            t_string = now.strftime("%d/%m/%Y %H:%M %p")
            requestedBook.objects.create(
                title=Book.title,
                author=Book.author,
                department=Book.department,
                shelf=Book.shelf,
                studentName=boarder.name,
                studentContact=boarder.contact,
                email=boarder.emails,
                created_at=t_string,
                status="requested",
                bookID=id,
            )
            x = Book.available
            Book.available = x - 1
            Book.save()
            boarder.save()
        else:
            messages.info(request, "You can only apply for 2 books at a time")
            return redirect("/library")
        messages.info(request, "Request submitted successfully")
        return redirect("/library")


def checkedOutBooks(request):
    if request.user.is_authenticated and request.user.email in allowedEmailsLibrary:
        now = datetime.now()
        books = requestedBook.objects.all()
        for Rbook in books:
            b = Rbook.created_at[0:10]
            a = datetime.strptime(b, "%d/%m/%Y")
            print((now - a).days)
            if (now - a).days > 3:
                boarder = azad_boarders.objects.get(emails=Rbook.email)
                boarder.books -= 1
                boarder.save()
                Book = book.objects.get(id=Rbook.bookID)
                Book.available += 1
                Book.save()
                Rbook.delete()
        requestedBooks = requestedBook.objects.filter(status="requested")
        requestedBooks_paginator = Paginator(requestedBooks, 10)
        current_page_requestedBooks = requestedBooks_paginator.page(
            request.GET.get("requestedBooks_page", 1)
        )

        checkedOutBooks = requestedBook.objects.filter(status="checkedOut")
        checkedOutBooks_paginator = Paginator(checkedOutBooks, 10)
        current_page_checkedOutBooks = checkedOutBooks_paginator.page(
            request.GET.get("checkedOutBooks_page", 1)
        )

        return render(
            request,
            "checkedOutBooks.html",
            {
                "requestedBooks": current_page_requestedBooks,
                "checkedOutBooks": current_page_checkedOutBooks,
            },
        )
    messages.info(request, "Please login with valid ID to access library records")
    return redirect("/")


def previousBookRequests(request):
    if request.user.is_authenticated:
        requestedBooks = requestedBook.objects.filter(
            status="requested", email=request.user.email
        )
        requestedBooks_paginator = Paginator(requestedBooks, 10)
        current_page_requestedBooks = requestedBooks_paginator.page(
            request.GET.get("requestedBooks_page", 1)
        )

        checkedOutBooks = requestedBook.objects.filter(
            status="checkedOut", email=request.user.email
        )
        checkedOutBooks_paginator = Paginator(checkedOutBooks, 10)
        current_page_checkedOutBooks = checkedOutBooks_paginator.page(
            request.GET.get("checkedOutBooks_page", 1)
        )

        return render(
            request,
            "previousBookRequests.html",
            {
                "requestedBooks": current_page_requestedBooks,
                "checkedOutBooks": current_page_checkedOutBooks,
            },
        )
    messages.info(request, "Please login with valid ID to access library records")
    return redirect("/")


def approve(request):
    if request.user.is_authenticated and request.user.email in allowedEmailsLibrary:
        if request.method == "POST":
            id = request.POST.get("id")
            Book = requestedBook.objects.get(id=id)
            Book.status = "checkedOut"
            now = datetime.now()
            t_string = now.strftime("%d/%m/%Y %H:%M %p")
            Book.created_at = t_string
            Book.save()
            return redirect("/checkedOutBooks")


def checkIn(request):
    if request.user.is_authenticated and request.user.email in allowedEmailsLibrary:
        if request.method == "POST":
            id = request.POST.get("id")
            RequestedBook = requestedBook.objects.get(id=id)
            Book = book.objects.get(id=RequestedBook.bookID)
            Book.available += 1
            Book.save()
            boarder = azad_boarders.objects.get(emails=RequestedBook.email)
            boarder.books -= 1
            boarder.save()
            RequestedBook.delete()
            return redirect("/checkedOutBooks")


def cancelBookRequest(request):
    id = request.POST.get("id")
    RequestedBook = requestedBook.objects.get(id=id)
    boarder = azad_boarders.objects.get(emails=RequestedBook.email)
    boarder.books -= 1
    boarder.save()
    Book = book.objects.get(id=RequestedBook.bookID)
    Book.available += 1
    Book.save()
    RequestedBook.delete()
    return redirect("/previousBookRequests")


def search(request):
    if request.method == "POST":
        str = request.POST.get("search")

        if str:
            books = book.objects.filter(
                models.Q(title__icontains=str)
                | models.Q(author__icontains=str)
                | models.Q(department__icontains=str)
            )
            if books:
                return library(request, books, str)
            else:
                books = book.objects.all()
                messages.info(request, "No books found")
                return redirect("/library")

    books = book.objects.all()
    return redirect("/library")


def about(request):
    if request.method == "POST":
        form = ContactForm(request.POST)
        form.save()
        subject = request.POST.get("subject", "")
        message = request.POST.get("msg", "")
        from_email = request.POST.get("email", "")
        if subject and message and from_email:
            try:
                send_mail(subject, message, from_email, ["smarakdev@gmail.com"])
            except BadHeaderError:
                return HttpResponse("Invalid header found.")
            return redirect("/about")
        else:
            form = ContactForm()
            return render(request, "about.html", {"form": form})
    else:
        form = ContactForm()
        return render(request, "about.html", {"form": form})


def is_ajax(request):
    return request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"


def postComment(request):
    # request should be ajax and method should be POST.
    print("hello")
    if is_ajax(request=request) and request.method == "POST":
        # get the form data
        form = CommentForm(request.POST)
        # save the data and after fetch the object in instance
        if form.is_valid():
            instance = form.save()
            # serialize in new friend object in json
            ser_instance = serializers.serialize(
                "json",
                [
                    instance,
                ],
            )
            # send to client side.
            print("saved")
            return JsonResponse({"instance": ser_instance}, status=200)
        else:
            # some form errors occured.
            return JsonResponse({"error": form.errors}, status=400)

    # some error occured
    return JsonResponse({"error": ""}, status=400)


def event(request, eventid):
    id = eventid
    # print("message", eventid)
    event = Event.objects.filter(id=eventid)
    itemlist = []
    # print(event)
    # print(event[0].para_set.all())
    commentform = CommentForm()
    comments = event[0].comments.all()

    cover = ""
    for item in event[0].imagemodel_set.all():
        if item.caption == "cover":
            cover = item
            continue
        itemlist.append(item)
    for item in event[0].para_set.all():
        itemlist.append(item)
    # print(itemlist)
    itemlist.sort(key=lambda x: x.date, reverse=True)
    # print(itemlist)
    # print(cover.caption)
    events = Event.objects.all()
    return render(
        request,
        "event.html",
        {
            "events": events,
            "cover": cover,
            "event": event[0],
            "form": commentform,
            "comments": comments,
        },
    )


def custom_logout(request):
    logout(request)
    message = "Logged out successfully"
    params = {"message": message}
    return render(request, "index.html", params)
